#!/usr/bin/env python3
"""Deja un Excel de diagnóstico de la corrida en `informes/`.

    python3 informe_corrida.py

Es el informe para MIRAR EL PROCESO, no el catastro. El sitio contesta "¿qué
hago hoy?"; esto contesta "¿está funcionando esto y dónde se está rompiendo?".
Por eso vive fuera de `web/` y fuera de git (`informes/` está en .gitignore):
es un cuaderno de trabajo, no un entregable.

Tres hojas, y la división no es decorativa:

  1. Diagnóstico — los números de la corrida y el delta contra la anterior.
     Sirve para ver de un vistazo si algo se cayó: una fuente que pasó de 200
     eventos a 0, la georreferenciación que bajó diez puntos, una corrida que
     se demoró el triple.

  2. Para revisar — la lista de eventos que el pipeline no supo resolver solo.
     Es una cola de trabajo: cada fila es una decisión que una persona puede
     tomar en diez segundos y que después se guarda en `config/correcciones/`
     para que no vuelva a preguntarse.

  3. Fuentes — el catastro: cada sitio del que se saca dato, con qué método se
     lee y qué pasó hoy con él. Sale del ARCHIVO DE CONFIGURACIÓN y no de la
     tabla `corridas`, porque la tabla sólo conoce a las fuentes que corrieron
     y acá la pregunta incluye a las que no: las apagadas, con la razón por la
     que se apagaron, y las encendidas que ni siquiera se intentaron. Cubre las
     dos mitades del pipeline —las fuentes de eventos y los bancos de
     descuentos— porque son sitios web ajenos igual de frágiles.

La comparación con "la corrida anterior" sale de `datos/historial_corridas.json`,
que este script escribe al final. Se usa eso y no `git show HEAD:web/eventos.json`
porque con `--sin-publicar` el HEAD no avanza y entonces todas las corridas del
día se comparaban contra el mismo punto, mostrando altas que ya se habían
mostrado. El historial guarda además una fila por corrida con los agregados,
así que a las pocas semanas el propio archivo muestra la tendencia.
"""

from __future__ import annotations

import json
import re
import sqlite3
import sys
from collections import Counter
from datetime import datetime, date
from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from loica.almacen import RUTA_DB, SQL_VIGENTE
from loica.clasificar import clasificar
from loica.modelo import es_url_publica

RAIZ = Path(__file__).resolve().parent
DIR_INFORMES = RAIZ / "informes"
RUTA_EVENTOS = RAIZ / "web" / "eventos.json"
RUTA_TALLERES = RAIZ / "web" / "talleres.json"
RUTA_HISTORIAL = RAIZ / "datos" / "historial_corridas.json"
RUTA_FUENTES = RAIZ / "config" / "fuentes.yaml"
RUTA_BANCOS = RAIZ / "config" / "bancos.yaml"
RUTA_ESTADO_DESCUENTOS = RAIZ / "datos" / "ultima_corrida_descuentos.json"
DIR_LOGS = RAIZ / "datos" / "logs"

# Las precisiones que significan "el pin está donde de verdad ocurre la cosa".
# `comuna` es el centro de la comuna y `sin_ubicar` es que no hay pin: las dos
# son deuda, y por eso se cuentan aparte.
PRECISIONES_EXACTAS = {"recinto", "fuente", "calle", "correccion"}

# Paleta del informe. Es la del sitio para que se reconozca de qué proyecto es,
# pero apagada: esto se lee en una planilla, no es una vitrina.
TINTA = "1E2A4A"
CREMA = "FAF3E7"
ARENA = "EDE7DE"
ROJO = "C0392B"
AMBAR = "B9770E"
VERDE = "1E7A52"

FUENTE = "Arial"
_borde = Side(style="thin", color="BEB6A9")
BORDE = Border(left=_borde, right=_borde, top=_borde, bottom=_borde)


# ---------------------------------------------------------------- datos

def _cargar_json(ruta: Path) -> dict:
    try:
        return json.loads(ruta.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return {}


def corrida_de_hoy(con: sqlite3.Connection) -> list[sqlite3.Row]:
    """Las filas de `corridas` de la última tanda, una por fuente.

    Se agrupa por fuente y se toma la más reciente: correr `run_diario.py` dos
    veces el mismo día dejaba dos filas por fuente y los totales salían al
    doble, que es justo el número que uno no quiere dudar en un diagnóstico.
    """
    return con.execute(
        """SELECT c.* FROM corridas c
           JOIN (SELECT fuente, MAX(momento) AS ultimo
                 FROM corridas WHERE date(momento) = date('now','localtime')
                 GROUP BY fuente) u
             ON c.fuente = u.fuente AND c.momento = u.ultimo
           ORDER BY c.fuente""",
    ).fetchall()


def eventos_publicados() -> list[dict]:
    """TODO lo publicado: panoramas y talleres juntos.

    El diagnóstico mira el proceso, y el proceso es uno solo — la extracción,
    la georreferenciación y la clasificación no distinguen formato. La
    separación en dos archivos es una decisión de las PÁGINAS, no del
    pipeline, así que acá se vuelven a juntar.
    """
    return (_cargar_json(RUTA_EVENTOS).get("eventos", [])
            + _cargar_json(RUTA_TALLERES).get("talleres", []))


def historial() -> dict:
    d = _cargar_json(RUTA_HISTORIAL)
    return {"ultima": d.get("ultima") or {}, "filas": d.get("filas") or []}


def lugares_nuevos(con: sqlite3.Connection) -> list[tuple[str, str, int]]:
    """Lugares que aparecen por primera vez en la base con esta corrida.

    Un lugar nuevo es donde se rompe la georreferenciación: el catastro no lo
    conoce, la memoria de correcciones tampoco, y el pin sale al centro de la
    comuna o no sale. Saber cuáles entraron hoy es saber qué revisar.
    """
    filas = con.execute(
        """SELECT lugar_nombre, comuna, COUNT(*) AS n
           FROM eventos
           WHERE lugar_nombre IS NOT NULL AND TRIM(lugar_nombre) != ''
           GROUP BY lugar_nombre
           HAVING MIN(date(fecha_extraccion)) = date('now','localtime')
           ORDER BY n DESC, lugar_nombre""",
    ).fetchall()
    return [(f["lugar_nombre"], f["comuna"] or "", f["n"]) for f in filas]


def clasificar_con_origen(con: sqlite3.Connection) -> dict[str, str]:
    """Para cada evento vigente, de dónde salió su categoría.

    `exportar_web.py` se queda solo con la categoría y bota el origen, porque
    al sitio no le sirve. Acá sí: una categoría sacada del título es un hecho
    leído y una sacada del recinto es una conjetura del 85%. Son la misma
    palabra en la tarjeta y no son la misma confianza, y esa diferencia es
    justamente lo que hay que poner en la cola de revisión.
    """
    origenes = {}
    for fila in con.execute(
            "SELECT hash_dedup, titulo, categoria, descripcion_corta, "
            "lugar_nombre, fuente_nombre FROM eventos WHERE " + SQL_VIGENTE):
        _, origen = clasificar(fila["titulo"], fila["categoria"] or "",
                               fila["descripcion_corta"] or "",
                               fila["lugar_nombre"] or "",
                               fila["fuente_nombre"] or "")
        origenes[fila["hash_dedup"]] = origen
    return origenes


# ------------------------------------------------------------- escritura

def _titulo(hoja, fila: int, texto: str, ancho: int = 6) -> int:
    c = hoja.cell(row=fila, column=1, value=texto)
    c.font = Font(name=FUENTE, size=12, bold=True, color=TINTA)
    c.fill = PatternFill("solid", fgColor=ARENA)
    for col in range(2, ancho + 1):
        hoja.cell(row=fila, column=col).fill = PatternFill("solid", fgColor=ARENA)
    return fila + 1


def _dato(hoja, fila: int, etiqueta: str, valor, nota: str = "",
          color: str | None = None, formato: str | None = None) -> int:
    a = hoja.cell(row=fila, column=1, value=etiqueta)
    a.font = Font(name=FUENTE, size=10, color=TINTA)
    b = hoja.cell(row=fila, column=2, value=valor)
    b.font = Font(name=FUENTE, size=10, bold=True, color=color or TINTA)
    b.alignment = Alignment(horizontal="right")
    if formato:
        b.number_format = formato
    if nota:
        c = hoja.cell(row=fila, column=3, value=nota)
        c.font = Font(name=FUENTE, size=9, italic=True, color="645D51")
    return fila + 1


def _encabezados(hoja, fila: int, titulos: list[str]) -> int:
    for i, t in enumerate(titulos, start=1):
        c = hoja.cell(row=fila, column=i, value=t)
        c.font = Font(name=FUENTE, size=9, bold=True, color=CREMA)
        c.fill = PatternFill("solid", fgColor=TINTA)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = BORDE
    hoja.row_dimensions[fila].height = 26
    return fila + 1


def hoja_diagnostico(wb: Workbook, ctx: dict) -> None:
    h = wb.active
    h.title = "Diagnóstico"
    h.sheet_view.showGridLines = False

    f = 1
    c = h.cell(row=1, column=1, value=f"Corrida del {ctx['momento']:%d-%m-%Y a las %H:%M}")
    c.font = Font(name=FUENTE, size=15, bold=True, color=TINTA)
    f = 3

    # ---- La corrida
    f = _titulo(h, f, "LA CORRIDA")
    f = _dato(h, f, "Duración total", ctx["duracion_min"],
              "minutos de punta a punta", formato="0.0")
    f = _dato(h, f, "Fuentes que corrieron", ctx["fuentes_total"])
    f = _dato(h, f, "Fuentes con error", ctx["fuentes_error"],
              "revísalas en la tabla de abajo" if ctx["fuentes_error"] else "ninguna",
              ROJO if ctx["fuentes_error"] else VERDE)
    f = _dato(h, f, "Fuentes vivas pero sin nada futuro", ctx["fuentes_vacias"],
              "responden bien y no aportan eventos: agenda abandonada o "
              "cambio de formato" if ctx["fuentes_vacias"] else "ninguna",
              AMBAR if ctx["fuentes_vacias"] else VERDE)
    # El fallo más silencioso de todos: la fuente que devuelve CERO sin error.
    # Passline estuvo así, pidiendo todos los días contra un 403 que el cliente
    # se tragaba, y en el informe se veía tan viva como las demás.
    f = _dato(h, f, "Fuentes que trajeron cero", ctx["fuentes_en_cero"],
              "sin error y sin eventos: es un 403 o un adaptador ciego, no "
              "una agenda tranquila" if ctx["fuentes_en_cero"] else "ninguna",
              ROJO if ctx["fuentes_en_cero"] else VERDE)
    f += 1

    # ---- Qué cambió
    f = _titulo(h, f, "QUÉ CAMBIÓ DESDE LA CORRIDA ANTERIOR")
    fila_hoy, fila_antes = f, f + 1
    f = _dato(h, f, "Publicados ahora", ctx["total_hoy"],
              f"{ctx['talleres_hoy']} talleres y clases, el resto panoramas")
    f = _dato(h, f, "Publicados en la corrida anterior", ctx["total_antes"],
              ctx["cuando_antes"])
    f = _dato(h, f, "Entraron (altas)", ctx["altas"], "no estaban antes", VERDE)
    f = _dato(h, f, "Salieron (bajas)", ctx["bajas"],
              "ya pasaron, caducaron o la fuente los bajó", AMBAR)
    fila_neto = f
    h.cell(row=f, column=1, value="Neto").font = Font(name=FUENTE, size=10, color=TINTA)
    nc = h.cell(row=f, column=2, value=f"=B{fila_hoy}-B{fila_antes}")
    nc.font = Font(name=FUENTE, size=10, bold=True, color=TINTA)
    nc.alignment = Alignment(horizontal="right")
    h.cell(row=f, column=3, value="si es muy negativo sin explicación, mira las "
           "fuentes con error").font = Font(name=FUENTE, size=9, italic=True, color="645D51")
    f += 2

    # ---- Extracción
    f = _titulo(h, f, "LO QUE TRAJO LA EXTRACCIÓN")
    f = _dato(h, f, "Eventos encontrados en las fuentes", ctx["encontrados"])
    f = _dato(h, f, "Nuevos guardados en la base", ctx["nuevos"])
    f = _dato(h, f, "Actualizados", ctx["actualizados"], "ya estaban y cambió algo")
    f = _dato(h, f, "Descartados en la puerta", ctx["descartados"],
              "sin título, sin link, pasados o link de máquina")
    f += 1

    # ---- Georreferenciación
    f = _titulo(h, f, "GEORREFERENCIACIÓN")
    fila_exactos = f
    f = _dato(h, f, "Con pin exacto", ctx["exactos"],
              "recinto, calle, coordenada de la fuente o corrección a mano", VERDE)
    f = _dato(h, f, "Pin aproximado (centro de comuna)", ctx["aprox"],
              "el mapa lo atenúa; están en la hoja 2", AMBAR)
    f = _dato(h, f, "Sin pin", ctx["sin_pin"],
              "salen en la lista pero no en el mapa; están en la hoja 2", ROJO)
    fila_total_geo = f
    f = _dato(h, f, "Total publicado", ctx["total_hoy"])
    pc = h.cell(row=f, column=1, value="% con pin exacto")
    pc.font = Font(name=FUENTE, size=10, color=TINTA)
    pv = h.cell(row=f, column=2, value=f"=IFERROR(B{fila_exactos}/B{fila_total_geo},0)")
    pv.font = Font(name=FUENTE, size=10, bold=True, color=TINTA)
    pv.number_format = "0.0%"
    pv.alignment = Alignment(horizontal="right")
    h.cell(row=f, column=3, value="es la métrica clave del mapa").font = Font(
        name=FUENTE, size=9, italic=True, color="645D51")
    f += 1
    f = _dato(h, f, "De los eventos nuevos de hoy, con pin", ctx["nuevos_con_pin"],
              f"de {ctx['nuevos_publicados']} nuevos que salieron publicados")
    f += 1

    # ---- Clasificación
    f = _titulo(h, f, "CLASIFICACIÓN")
    f = _dato(h, f, "Leída del título o la descripción", ctx["origen_texto"],
              "el clasificador lo leyó, no lo adivinó", VERDE)
    f = _dato(h, f, "Decidida por la memoria de categorías", ctx["origen_memoria"],
              "una regla de config/correcciones/categorias.yaml", VERDE)
    f = _dato(h, f, "Adivinada por el recinto o la fuente", ctx["origen_prior"],
              "conjetura razonable (~85% de acierto); están en la hoja 2", AMBAR)
    f = _dato(h, f, "Sin señal: cayó en «otros»", ctx["origen_defecto"],
              "el clasificador no supo qué es; están en la hoja 2", ROJO)
    f = _dato(h, f, "Con subcategoría", ctx["con_subcat"],
              "género de la fiesta, tipo de obra, oficio del taller")
    f += 1
    f = _encabezados(h, f, ["Categoría", "Eventos", "% del total"])
    fila_cat = f
    for cat, n in ctx["por_categoria"]:
        h.cell(row=f, column=1, value=cat).font = Font(name=FUENTE, size=10)
        h.cell(row=f, column=2, value=n).font = Font(name=FUENTE, size=10)
        p = h.cell(row=f, column=3, value=f"=IFERROR(B{f}/$B${fila_total_geo},0)")
        p.font = Font(name=FUENTE, size=10)
        p.number_format = "0.0%"
        for col in (1, 2, 3):
            h.cell(row=f, column=col).border = BORDE
        f += 1
    tc = h.cell(row=f, column=1, value="Total")
    tc.font = Font(name=FUENTE, size=10, bold=True)
    tv = h.cell(row=f, column=2, value=f"=SUM(B{fila_cat}:B{f - 1})")
    tv.font = Font(name=FUENTE, size=10, bold=True)
    for col in (1, 2, 3):
        h.cell(row=f, column=col).border = BORDE
    f += 2

    # ---- Lugares nuevos
    f = _titulo(h, f, f"LUGARES QUE APARECEN POR PRIMERA VEZ ({len(ctx['lugares'])})")
    h.cell(row=f, column=1, value="Un lugar nuevo es donde se rompe la "
           "georreferenciación: nadie lo ha ubicado todavía.").font = Font(
        name=FUENTE, size=9, italic=True, color="645D51")
    f += 1
    if ctx["lugares"]:
        f = _encabezados(h, f, ["Lugar", "Comuna", "Eventos"])
        for nombre, comuna, n in ctx["lugares"][:60]:
            h.cell(row=f, column=1, value=nombre).font = Font(name=FUENTE, size=10)
            h.cell(row=f, column=2, value=comuna).font = Font(name=FUENTE, size=10)
            h.cell(row=f, column=3, value=n).font = Font(name=FUENTE, size=10)
            for col in (1, 2, 3):
                h.cell(row=f, column=col).border = BORDE
            f += 1
        if len(ctx["lugares"]) > 60:
            h.cell(row=f, column=1,
                   value=f"… y {len(ctx['lugares']) - 60} más").font = Font(
                name=FUENTE, size=9, italic=True, color="645D51")
            f += 1
    else:
        h.cell(row=f, column=1, value="Ninguno: todas las sedes de hoy ya se "
               "conocían.").font = Font(name=FUENTE, size=10, italic=True)
        f += 1
    f += 1

    # ---- Por fuente
    f = _titulo(h, f, "POR FUENTE")
    f = _encabezados(h, f, ["Fuente", "Encontrados", "Nuevos", "Actualizados",
                            "Descartados", "Segundos", "Estado"])
    fila_fuentes = f
    for fila in ctx["fuentes"]:
        estado = "ERROR" if fila["error"] else ("sin nada futuro"
                 if fila["fuente"] in ctx["nombres_vacias"] else "ok")
        color = ROJO if fila["error"] else (AMBAR if estado != "ok" else TINTA)
        valores = [fila["fuente"], fila["encontrados"], fila["nuevos"],
                   fila["actualizados"], fila["descartados"],
                   round(fila["duracion_seg"] or 0, 1), estado]
        for i, v in enumerate(valores, start=1):
            c = h.cell(row=f, column=i, value=v)
            c.font = Font(name=FUENTE, size=10,
                          color=color if i in (1, 7) else TINTA)
            c.border = BORDE
        if fila["error"]:
            h.cell(row=f, column=7).comment = None
            h.cell(row=f, column=7).value = f"ERROR: {str(fila['error'])[:180]}"
        f += 1
    if ctx["fuentes"]:
        h.cell(row=f, column=1, value="Total").font = Font(name=FUENTE, size=10, bold=True)
        for col in range(2, 7):
            L = get_column_letter(col)
            t = h.cell(row=f, column=col,
                       value=f"=SUM({L}{fila_fuentes}:{L}{f - 1})")
            t.font = Font(name=FUENTE, size=10, bold=True)
            t.border = BORDE
        h.cell(row=f, column=1).border = BORDE
        h.cell(row=f, column=7).border = BORDE

    for col, ancho in zip("ABCDEFG", (46, 13, 52, 13, 13, 11, 20)):
        h.column_dimensions[col].width = ancho


def hoja_revisar(wb: Workbook, ctx: dict) -> None:
    h = wb.create_sheet("Para revisar")
    h.sheet_view.showGridLines = False

    c = h.cell(row=1, column=1, value="Lo que el pipeline no supo resolver solo")
    c.font = Font(name=FUENTE, size=15, bold=True, color=TINTA)
    h.cell(row=2, column=1, value=(
        "Cada fila es una decisión de diez segundos. Lo que se arregle acá va a "
        "config/correcciones/ (lugares.yaml para la ubicación, eventos.yaml para "
        "la categoría) y la corrida siguiente lo aplica sola. Si el mismo error "
        "se repite en eventos nuevos, el arreglo va en loica/clasificar.py, no "
        "en la memoria.")).font = Font(name=FUENTE, size=9, italic=True, color="645D51")
    h.merge_cells(start_row=2, start_column=1, end_row=2, end_column=11)
    h.row_dimensions[2].height = 30
    h.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    columnas = ["Qué revisar", "Título", "Cuándo", "Lugar", "Dirección",
                "Comuna", "Categoría", "De dónde salió", "Subcategoría",
                "Fuente", "Link"]
    f = _encabezados(h, 4, columnas)
    h.freeze_panes = h.cell(row=f, column=1)

    for e in ctx["revisar"]:
        valores = [e["motivo"], e["titulo"], e["cuando"], e["lugar"],
                   e["direccion"], e["comuna"], e["categoria"], e["origen"],
                   e["subcategoria"], e["fuente"], e["url"]]
        for i, v in enumerate(valores, start=1):
            c = h.cell(row=f, column=i, value=v)
            c.font = Font(name=FUENTE, size=10, color=TINTA)
            c.border = BORDE
            c.alignment = Alignment(vertical="top", wrap_text=(i in (2, 4, 5)))
        h.cell(row=f, column=1).font = Font(name=FUENTE, size=10, bold=True,
                                            color=e["color"])
        if e["url"]:
            enlace = h.cell(row=f, column=11)
            enlace.hyperlink = e["url"]
            enlace.value = "abrir"
            enlace.font = Font(name=FUENTE, size=10, color="1B6FD1", underline="single")
        f += 1

    h.auto_filter.ref = f"A4:K{max(f - 1, 5)}"
    for col, ancho in zip("ABCDEFGHIJK", (30, 46, 17, 30, 30, 16, 12, 16, 15, 22, 9)):
        h.column_dimensions[col].width = ancho


def _leer_yaml(ruta: Path, clave: str) -> list[dict]:
    """El catálogo de fuentes tal como está escrito en config/.

    Si el archivo no se puede leer se devuelve vacío en vez de reventar: el
    diagnóstico no bloquea la corrida, y una hoja de fuentes incompleta es
    mejor que ninguna.
    """
    try:
        import yaml
        with open(ruta, encoding="utf-8") as f:
            return yaml.safe_load(f).get(clave) or []
    except Exception:
        return []


# Cómo se saca el dato de cada tipo de fuente, dicho en castellano. El YAML
# guarda el nombre del adaptador (`wordpress`, `sitemap`) porque es lo que el
# código necesita; acá interesa poder contestar "¿de dónde sale esto?" sin
# abrir el repositorio.
METODOS_EVENTOS = {
    "wordpress": "API REST de WordPress (/wp-json)",
    "eventon": "Calendario EventON (admin-ajax)",
    "rss": "Feed RSS",
    "html": "HTML de la agenda con selectores",
    "sitemap": "Sitemap XML + ficha de cada evento",
    "carteleras": "Índice de carteleras + ficha por local",
    "cine": "Cartelera de cine (una fila por función)",
    "tabla": "Tabla HTML de la propia página",
    "json": "API JSON, mapeo declarado en el YAML",
    "manual": "Captura a mano (datos/manual)",
    "ticketmaster": "API Discovery de Ticketmaster (con key)",
}

METODOS_BANCOS = {
    "bancochile": "CMS público de beneficios (JSON)",
    "bci": "Portal vivirconbeneficios (JSON por categoría)",
    "falabella": "Contentful Content Delivery (JSON)",
    "cencosud": "JSON incrustado en la página (window.CardsAPI)",
    "santander": "Captura a mano: el sitio bloquea el rastreo",
}

# El orden en que se leen las filas. Primero lo que hay que arreglar, al final
# lo que está bien y lo que está apagado a propósito.
ORDEN_ESTADOS = {
    "error": 0,
    "no corrió": 1,
    "trajo cero": 2,
    "sin nada futuro": 3,
    "extrajo": 4,
    "apagada": 5,
}
COLOR_ESTADOS = {
    "error": ROJO,
    "no corrió": ROJO,
    "trajo cero": ROJO,
    "sin nada futuro": AMBAR,
    "extrajo": VERDE,
    "apagada": "645D51",
}


def _resumir_nota(texto: str, tope: int = 220) -> str:
    """La nota del YAML resumida: sirve para saber por qué una fuente está apagada."""
    limpio = " ".join((texto or "").split())
    return limpio[:tope] + ("…" if len(limpio) > tope else "")


_HORA = re.compile(r"^\d\d:\d\d:\d\d\s")
_FALLO_RED = re.compile(
    r"^(\d\d:\d\d:\d\d)\s+\w+\s+loica\.red\s+"
    r"(?:Falló (\S+): (.*)|robots\.txt prohíbe (\S+).*)$")


# Los errores de red que se repiten, dichos como los diría una persona. El
# texto de requests es exacto y también es una pared: "HTTPSConnectionPool(
# host='parquemet.cl', port=443): Max retries exceeded... NewConnectionError"
# son tres capas de librería para decir que el dominio no existe. En una
# planilla que se lee en diez segundos, eso vale menos que una frase.
CAUSAS = (
    ("nodename nor servname", "El dominio no resuelve en DNS: el sitio no "
                              "existe hoy para nadie, no es cosa nuestra."),
    ("Name or service not known", "El dominio no resuelve en DNS: el sitio no "
                                  "existe hoy para nadie, no es cosa nuestra."),
    ("Read timed out", "El sitio aceptó la conexión y no contestó a tiempo. "
                       "Suele ser lentitud del servidor, no un bloqueo."),
    ("Connection timed out", "El sitio no aceptó la conexión: cortafuegos o "
                             "servidor caído."),
    ("Connection refused", "El servidor rechazó la conexión."),
    ("SSLError", "El certificado del sitio está roto o vencido."),
    ("CertificateError", "El certificado del sitio está roto o vencido."),
    ("robots.txt", "robots.txt lo prohíbe: no se rastrea a propósito."),
)


def _en_castellano(mensaje: str) -> str:
    for marca, explicacion in CAUSAS:
        if marca in mensaje:
            return explicacion
    return mensaje[:200]


def fallos_de_red(desde: str) -> dict[str, dict]:
    """Los errores HTTP de esta corrida, agrupados por dominio.

    Existe por el fallo más caro de leer que tiene el pipeline: una fuente a la
    que se le cayeron TODAS las peticiones no queda registrada como error.
    `ClienteEducado.obtener` devuelve None cuando la red falla —para que una
    URL rota no tumbe la corrida entera— y el adaptador sigue con la
    siguiente, así que la fuente termina con `error = NULL` y `encontrados = 0`.
    En la tabla `corridas` eso es indistinguible de una agenda vacía, y son
    cosas opuestas: una hay que arreglarla hoy y la otra no es problema.

    El motivo real sí quedó escrito, en `datos/logs/`. Acá se lee de vuelta y
    se le devuelve a la fuente que le corresponde, cruzando por dominio.

    El log guarda la hora y no la fecha, así que hay que acotar a mano qué
    tramo es esta corrida: se recorre al revés y se corta en la primera línea
    ANTERIOR a `desde`, que es la hora en que arrancó la primera fuente (sale
    de `corridas.momento`, que sí trae fecha). Todo lo de más arriba es de una
    corrida previa —de hoy más temprano o de ayer— y contarlo sería atribuirle
    a esta corrida caídas que ya pasaron.

    El corte NO puede ser "ahora": mientras se lee el archivo puede seguir
    creciendo, y entonces las líneas más nuevas que ese instante cortarían el
    escaneo antes de tiempo, perdiendo justo las caídas de la corrida en curso.
    """
    archivo = DIR_LOGS / f"{datetime.now():%Y-%m}.log"
    try:
        lineas = archivo.read_text(encoding="utf-8", errors="replace").splitlines()
    except OSError:
        return {}

    por_dominio: dict[str, dict] = {}
    for linea in reversed(lineas):
        # El corte se mira en TODA línea con hora, no sólo en las que fallaron:
        # si no, un mes sin caídas obliga a recorrer el archivo entero para no
        # encontrar nada.
        if _HORA.match(linea) and linea[:8] < desde:
            break
        coincidencia = _FALLO_RED.match(linea)
        if not coincidencia:
            continue
        _, url_fallo, mensaje, url_robots = coincidencia.groups()
        url = url_fallo or url_robots
        motivo = mensaje if url_fallo else "robots.txt lo prohíbe"
        dominio = urlparse(url).netloc.lower().removeprefix("www.")
        if not dominio:
            continue
        registro = por_dominio.setdefault(dominio, {"n": 0, "motivo": ""})
        registro["n"] += 1
        # Se recorre al revés, así que la última que se escribe es la primera
        # que ocurrió: la causa raíz, no la consecuencia.
        registro["motivo"] = _en_castellano(" ".join(motivo.split()))
    return por_dominio


def _caida_de(fuente: dict, caidas: dict[str, dict]) -> dict | None:
    """Los fallos de red que le tocan a esta fuente, buscados por su dominio.

    Se cruza por dominio y no por URL exacta porque una fuente pide varios
    endpoints (WordPress prueba un tipo de post tras otro) y lo que interesa
    no es cuál falló sino que el sitio no está contestando.
    """
    dominio = urlparse(fuente.get("url_base") or "").netloc.lower()
    return caidas.get(dominio.removeprefix("www.")) if dominio else None


def catastro_fuentes(ctx: dict) -> tuple[list[dict], list[dict]]:
    """Una fila por fuente del catálogo, corriera o no.

    Se parte del ARCHIVO DE CONFIGURACIÓN y no de la tabla `corridas`, y esa
    es toda la diferencia. La tabla sólo sabe de las fuentes que efectivamente
    corrieron: leyéndola se ve muy bien cuál falló, y no se ve en absoluto cuál
    ni siquiera se intentó. Para un catastro —que es la pregunta "¿qué estamos
    mirando de la web y qué no?"— justamente las que faltan son el dato.
    """
    corridas = {f["fuente"]: f for f in ctx["fuentes"]}
    en_sitio = Counter(e.get("fuente") or "" for e in ctx["publicados"])
    # Cuándo arrancó la primera fuente: es el borde inferior del tramo de log
    # que le pertenece a esta corrida. Si no hay corridas registradas no se
    # acota nada y se devuelve vacío, que es mejor que atribuir mal.
    momentos = [f["momento"] for f in ctx["fuentes"] if f["momento"]]
    caidas = fallos_de_red(min(momentos)[11:19]) if momentos else {}

    filas = []
    for fuente in _leer_yaml(RUTA_FUENTES, "fuentes"):
        nombre = fuente.get("nombre", "")
        activa = bool(fuente.get("activa", True))
        corrida = corridas.get(nombre)
        publicados = en_sitio.get(nombre, 0)

        if not activa:
            estado, detalle = "apagada", _resumir_nota(fuente.get("notas", ""))
            if publicados:
                detalle = (f"Apagada, y {publicados} eventos vigentes en el "
                           "sitio llevan su nombre: o quedaron de cuando estaba "
                           "encendida, o entraron por la ingesta asistida, que "
                           "conserva el nombre de la fuente original. ") + detalle
        elif corrida is None:
            estado = "no corrió"
            detalle = ("Está encendida y no tiene corrida de hoy: o se corrió "
                       "con --fuente, o la corrida se cortó antes de llegar.")
        elif corrida["error"]:
            estado, detalle = "error", str(corrida["error"])[:400]
        elif not (corrida["encontrados"] or 0):
            estado = "trajo cero"
            caida = _caida_de(fuente, caidas)
            detalle = (
                f"No es una agenda vacía: se le cayeron {caida['n']} peticiones. "
                f"{caida['motivo']}" if caida else
                "Respondió sin error y no trajo nada. Casi siempre es un 403 "
                "que el cliente se traga o un adaptador que quedó ciego porque "
                "el sitio cambió de formato.")
        elif fuente.get("tipo_adaptador") == "manual":
            # La ingesta asistida es la única fuente que no publica bajo su
            # propio nombre: cada evento capturado a mano conserva el de su
            # origen (Passline, un afiche, una cuenta de Instagram), que es
            # lo correcto para la atribución. El costo es que sus eventos
            # nunca aparecen contados a su nombre, y sin esta excepción la
            # fuente que más aporta salía marcada como "sin nada futuro".
            estado = "extrajo"
            detalle = ("Sus eventos se publican con el nombre de la fuente "
                       "original, no con el suyo: por eso «En el sitio» va en "
                       "cero. Se cuentan en las filas de esas fuentes.")
        elif nombre in ctx["nombres_vacias"]:
            estado = "sin nada futuro"
            caida = _caida_de(fuente, caidas)
            detalle = ("Trae eventos pero ninguno vigente: agenda abandonada o "
                       "fechas que el pipeline no supo leer.")
            if caida:
                detalle += (f" Ojo: además se le cayeron {caida['n']} "
                            f"peticiones. {caida['motivo']}")
        else:
            estado, detalle = "extrajo", ""

        filas.append({
            "nombre": nombre,
            "comuna": fuente.get("comuna", ""),
            "metodo": METODOS_EVENTOS.get(fuente.get("tipo_adaptador", ""),
                                          fuente.get("tipo_adaptador", "")),
            "url": fuente.get("url_agenda") or fuente.get("url_base") or "",
            "activa": "sí" if activa else "no",
            "estado": estado,
            "encontrados": (corrida["encontrados"] or 0) if corrida else "",
            "nuevos": (corrida["nuevos"] or 0) if corrida else "",
            "actualizados": (corrida["actualizados"] or 0) if corrida else "",
            "descartados": (corrida["descartados"] or 0) if corrida else "",
            "en_sitio": publicados,
            "segundos": round(corrida["duracion_seg"] or 0, 1) if corrida else "",
            "detalle": detalle,
        })

    estado_bancos = _cargar_json(RUTA_ESTADO_DESCUENTOS)
    corridos = {b["banco"]: b for b in (estado_bancos.get("bancos") or [])}

    bancos = []
    for banco in _leer_yaml(RUTA_BANCOS, "bancos"):
        nombre = banco.get("nombre", "")
        activo = bool(banco.get("activo", True))
        c = corridos.get(nombre)

        if not activo:
            estado, detalle = "apagada", _resumir_nota(banco.get("notas", ""))
        elif c is None:
            estado = "no corrió"
            detalle = ("Sin registro de la corrida de descuentos. Los descuentos "
                       "no abortan la corrida cuando fallan: puede haberse caído "
                       "el paso entero.")
        elif c["error"]:
            estado, detalle = "error", str(c["error"])[:400]
        elif not c["crudos"]:
            estado, detalle = "trajo cero", "Respondió sin error y no trajo nada."
        elif not c["vigentes"]:
            estado = "sin nada futuro"
            detalle = (f"Trajo {c['crudos']} promociones y ninguna quedó: "
                       f"{c['vencidos']} vencidas, {c['fuera_rm']} fuera de la RM.")
        else:
            estado, detalle = "extrajo", ""
        if c and banco.get("archivo"):
            detalle = (detalle + " Fuente de captura manual: no se rastrea, "
                       f"se anota en {banco['archivo']}.").strip()

        bancos.append({
            "nombre": nombre,
            "comuna": banco.get("emisor", ""),
            "metodo": METODOS_BANCOS.get(banco.get("adaptador", ""),
                                         banco.get("adaptador", "")),
            "url": banco.get("url_agenda") or banco.get("url_base")
                   or banco.get("archivo", ""),
            "activa": "sí" if activo else "no",
            "estado": estado,
            "encontrados": c["crudos"] if c else "",
            "nuevos": c["vigentes"] if c else "",
            "actualizados": c["con_dia"] if c else "",
            "descartados": c["vencidos"] if c else "",
            "en_sitio": c["vigentes"] if c else "",
            "segundos": "",
            "detalle": detalle,
        })

    clave = lambda f: (ORDEN_ESTADOS.get(f["estado"], 9), f["nombre"].lower())
    return sorted(filas, key=clave), sorted(bancos, key=clave)


def hoja_fuentes(wb: Workbook, ctx: dict) -> None:
    """El catastro: de dónde sale cada dato y qué pasó hoy con cada fuente.

    La hoja 1 dice si la corrida salió bien y la hoja 2 qué hay que arreglar a
    mano. Ésta contesta la pregunta de más atrás: qué sitios está mirando el
    proyecto, cómo los lee, y cuáles de ellos aportaron algo hoy y cuáles no.
    """
    h = wb.create_sheet("Fuentes")
    h.sheet_view.showGridLines = False

    eventos, bancos = ctx["catastro"], ctx["catastro_bancos"]
    todas = eventos + bancos
    encendidas = [f for f in todas if f["activa"] == "sí"]
    extrajeron = [f for f in todas if f["estado"] == "extrajo"]
    con_problema = [f for f in encendidas if f["estado"] in
                    ("error", "no corrió", "trajo cero", "sin nada futuro")]

    c = h.cell(row=1, column=1, value="De dónde sale cada dato")
    c.font = Font(name=FUENTE, size=15, bold=True, color=TINTA)
    h.cell(row=2, column=1, value=(
        "Todas las fuentes del catálogo, su método de extracción y qué pasó con "
        "cada una en esta corrida. Las que están encendidas y no aportaron nada "
        "van arriba de todo: ésas son las que hay que mirar. Las apagadas quedan "
        "al final con la razón por la que se apagaron — el catastro incluye lo "
        "que NO se está extrayendo, que es la mitad del dato."
    )).font = Font(name=FUENTE, size=9, italic=True, color="645D51")
    h.merge_cells(start_row=2, start_column=1, end_row=2, end_column=13)
    h.row_dimensions[2].height = 30
    h.cell(row=2, column=1).alignment = Alignment(wrap_text=True, vertical="top")

    f = 4
    f = _titulo(h, f, "EL CATASTRO EN UNA LÍNEA", ancho=13)
    f = _dato(h, f, "Fuentes en el catálogo", len(todas),
              f"{len(eventos)} de eventos y {len(bancos)} de descuentos")
    f = _dato(h, f, "Encendidas", len(encendidas),
              f"{len(todas) - len(encendidas)} apagadas a propósito")
    f = _dato(h, f, "Extrajeron datos hoy", len(extrajeron),
              "trajeron algo y tienen eventos vigentes", VERDE)
    f = _dato(h, f, "Encendidas que no aportaron nada", len(con_problema),
              "error, cero, o vivas sin nada futuro: la lista está abajo",
              ROJO if con_problema else VERDE)
    f += 1

    columnas = ["Fuente", "Comuna", "Método de extracción", "URL", "Encendida",
                "Estado", "Encontrados", "Nuevos", "Actualizados", "Descartados",
                "En el sitio", "Segundos", "Qué pasó"]
    f = _titulo(h, f, f"EVENTOS ({len(eventos)} fuentes)", ancho=13)
    f = _encabezados(h, f, columnas)
    primera_tabla = f
    h.freeze_panes = h.cell(row=f, column=1)
    f = _filas_fuentes(h, f, eventos)
    f += 1

    columnas_banco = ["Banco", "Emisor", "Método de extracción", "URL", "Encendido",
                      "Estado", "Promociones", "Vigentes en la RM", "Con día",
                      "Vencidas", "En el sitio", "", "Qué pasó"]
    f = _titulo(h, f, f"DESCUENTOS ({len(bancos)} bancos)", ancho=13)
    f = _encabezados(h, f, columnas_banco)
    f = _filas_fuentes(h, f, bancos)

    # El autofiltro va sobre la tabla de eventos, que es la larga: son 127 filas
    # y la pregunta habitual ("muéstrame sólo las de WordPress que fallaron") no
    # se contesta de otra forma.
    h.auto_filter.ref = f"A{primera_tabla - 1}:M{primera_tabla + len(eventos) - 1}"
    for col, ancho in zip(["A", "B", "C", "D", "E", "F", "G", "H", "I", "J",
                           "K", "L", "M"],
                          (40, 20, 36, 50, 11, 16, 12, 9, 12, 12, 11, 10, 70)):
        h.column_dimensions[col].width = ancho


def _filas_fuentes(hoja, fila: int, filas: list[dict]) -> int:
    for r in filas:
        valores = [r["nombre"], r["comuna"], r["metodo"], r["url"], r["activa"],
                   r["estado"], r["encontrados"], r["nuevos"], r["actualizados"],
                   r["descartados"], r["en_sitio"], r["segundos"], r["detalle"]]
        for i, v in enumerate(valores, start=1):
            c = hoja.cell(row=fila, column=i, value=v)
            c.font = Font(name=FUENTE, size=10, color=TINTA)
            c.border = BORDE
            c.alignment = Alignment(vertical="top", wrap_text=(i in (1, 3, 13)))
        hoja.cell(row=fila, column=6).font = Font(
            name=FUENTE, size=10, bold=True,
            color=COLOR_ESTADOS.get(r["estado"], TINTA))
        # La URL va como enlace sólo si es http(s): en el catálogo hay fuentes
        # de captura manual cuyo "origen" es una ruta de archivo, y un
        # hipervínculo a eso no lleva a ninguna parte.
        if es_url_publica(r["url"]):
            enlace = hoja.cell(row=fila, column=4)
            enlace.hyperlink = r["url"]
            enlace.font = Font(name=FUENTE, size=10, color="1B6FD1",
                               underline="single")
        fila += 1
    return fila

# ------------------------------------------------------------------ main

def armar_contexto() -> dict:
    con = sqlite3.connect(RUTA_DB)
    con.row_factory = sqlite3.Row

    fuentes = corrida_de_hoy(con)
    publicados = eventos_publicados()
    hist = historial()
    ids_antes = set(hist["ultima"].get("ids") or [])
    ids_hoy = {e["id"] for e in publicados}
    origenes = clasificar_con_origen(con)

    # Qué eventos entraron hoy a la base: sirve para saber si lo nuevo llegó
    # bien ubicado o si la deuda de georreferenciación la generan los recién
    # llegados.
    nuevos_hoy = {f["hash_dedup"] for f in con.execute(
        "SELECT hash_dedup FROM eventos "
        "WHERE date(fecha_extraccion) = date('now','localtime')")}
    pub_nuevos = [e for e in publicados if e["id"] in nuevos_hoy]

    nombres_vacias = set()
    for fila in fuentes:
        if fila["error"] or not fila["encontrados"]:
            continue
        vig = con.execute("SELECT COUNT(*) FROM eventos WHERE fuente_nombre = ? AND "
                          + SQL_VIGENTE, (fila["fuente"],)).fetchone()[0]
        if not vig:
            nombres_vacias.add(fila["fuente"])

    conteo_origen = Counter(origenes.get(e["id"], "defecto") for e in publicados)

    # ---- La cola de revisión
    ETIQUETAS = {
        "sin_pin": ("1. Sin pin en el mapa", ROJO),
        "sin_categoria": ("2. Sin categoría (otros)", ROJO),
        "categoria_adivinada": ("3. Categoría adivinada por el recinto", AMBAR),
        "pin_aproximado": ("4. Pin al centro de la comuna", AMBAR),
    }
    revisar = []
    for e in publicados:
        origen = origenes.get(e["id"], "")
        if e["precision"] == "sin_ubicar":
            motivo = "sin_pin"
        elif e["categoria"] == "otros":
            motivo = "sin_categoria"
        elif origen == "prior":
            motivo = "categoria_adivinada"
        elif e["precision"] == "comuna":
            motivo = "pin_aproximado"
        else:
            continue
        etiqueta, color = ETIQUETAS[motivo]
        revisar.append({
            "orden": etiqueta, "motivo": etiqueta, "color": color,
            "titulo": e["titulo"],
            "cuando": (e["inicio"] or "")[:16].replace("T", " "),
            "lugar": e["lugar"], "direccion": e["direccion"],
            "comuna": e["comuna"], "categoria": e["categoria"],
            "origen": origen, "subcategoria": e.get("subcategoria") or "",
            "fuente": e["fuente"], "url": e["url"],
        })
    revisar.sort(key=lambda r: (r["orden"], r["cuando"]))

    duracion = sum(f["duracion_seg"] or 0 for f in fuentes)
    exactos = sum(1 for e in publicados if e["precision"] in PRECISIONES_EXACTAS)

    ctx = {
        "momento": datetime.now(),
        "duracion_min": round(duracion / 60, 1),
        "fuentes": fuentes,
        "fuentes_total": len(fuentes),
        "fuentes_error": sum(1 for f in fuentes if f["error"]),
        "fuentes_vacias": len(nombres_vacias),
        "fuentes_en_cero": sum(1 for f in fuentes
                               if not f["error"] and not (f["encontrados"] or 0)),
        "nombres_vacias": nombres_vacias,
        "encontrados": sum(f["encontrados"] or 0 for f in fuentes),
        "nuevos": sum(f["nuevos"] or 0 for f in fuentes),
        "actualizados": sum(f["actualizados"] or 0 for f in fuentes),
        "descartados": sum(f["descartados"] or 0 for f in fuentes),
        "total_hoy": len(publicados),
        "talleres_hoy": sum(1 for e in publicados if e.get("formato") == "taller"),
        "total_antes": hist["ultima"].get("total", 0),
        "cuando_antes": hist["ultima"].get("momento", "primera corrida registrada"),
        "altas": len(ids_hoy - ids_antes) if ids_antes else len(ids_hoy),
        "bajas": len(ids_antes - ids_hoy),
        "exactos": exactos,
        "aprox": sum(1 for e in publicados if e["precision"] == "comuna"),
        "sin_pin": sum(1 for e in publicados if e["precision"] == "sin_ubicar"),
        "nuevos_publicados": len(pub_nuevos),
        "nuevos_con_pin": sum(1 for e in pub_nuevos if e["lat"] is not None),
        "origen_texto": conteo_origen["titulo"] + conteo_origen["etiqueta"] + conteo_origen["descripcion"],
        "origen_memoria": conteo_origen["memoria"],
        "origen_prior": conteo_origen["prior"],
        "origen_defecto": conteo_origen["defecto"],
        "con_subcat": sum(1 for e in publicados if e.get("subcategoria")),
        "por_categoria": Counter(e["categoria"] for e in publicados).most_common(),
        "lugares": lugares_nuevos(con),
        "revisar": revisar,
        "publicados": publicados,
    }
    ctx["catastro"], ctx["catastro_bancos"] = catastro_fuentes(ctx)
    con.close()
    return ctx, ids_hoy


def guardar_historial(ctx: dict, ids_hoy: set) -> None:
    hist = historial()
    hist["filas"].append({
        "momento": ctx["momento"].isoformat(timespec="seconds"),
        "total": ctx["total_hoy"], "altas": ctx["altas"], "bajas": ctx["bajas"],
        "exactos": ctx["exactos"], "sin_pin": ctx["sin_pin"],
        "otros": dict(ctx["por_categoria"]).get("otros", 0),
        "duracion_min": ctx["duracion_min"],
        "fuentes_error": ctx["fuentes_error"],
    })
    # 400 corridas son más de un año a una por día: pasado eso el archivo
    # crece sin que nadie lo lea.
    hist["filas"] = hist["filas"][-400:]
    hist["ultima"] = {
        "momento": ctx["momento"].strftime("%d-%m-%Y %H:%M"),
        "total": ctx["total_hoy"],
        "ids": sorted(ids_hoy),
    }
    RUTA_HISTORIAL.parent.mkdir(parents=True, exist_ok=True)
    RUTA_HISTORIAL.write_text(json.dumps(hist, ensure_ascii=False), encoding="utf-8")


def main() -> int:
    if not RUTA_EVENTOS.exists():
        print("No existe web/eventos.json: corre primero exportar_web.py")
        return 1

    ctx, ids_hoy = armar_contexto()

    wb = Workbook()
    hoja_diagnostico(wb, ctx)
    hoja_revisar(wb, ctx)
    hoja_fuentes(wb, ctx)

    DIR_INFORMES.mkdir(parents=True, exist_ok=True)
    ruta = DIR_INFORMES / f"{date.today():%Y-%m-%d}_diagnostico.xlsx"
    wb.save(ruta)
    guardar_historial(ctx, ids_hoy)

    print(f"Diagnóstico: {ruta}")
    print(f"  {ctx['total_hoy']} publicados "
          f"(+{ctx['altas']} / -{ctx['bajas']} desde la corrida anterior)")
    print(f"  {ctx['exactos']} con pin exacto, {ctx['sin_pin']} sin pin")
    print(f"  {len(ctx['revisar'])} eventos en la cola de revisión")
    catastro = ctx["catastro"] + ctx["catastro_bancos"]
    encendidas = [f for f in catastro if f["activa"] == "sí"]
    print(f"  {sum(1 for f in encendidas if f['estado'] == 'extrajo')} de "
          f"{len(encendidas)} fuentes encendidas extrajeron datos "
          f"(catálogo completo: {len(catastro)})")
    return 0


if __name__ == "__main__":
    sys.exit(main())
